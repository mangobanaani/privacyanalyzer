"""Excel spreadsheet processing."""

from pathlib import Path
from typing import List, Tuple
import openpyxl

from src.utils import get_logger

logger = get_logger(__name__)


class ExcelProcessor:
    """Extract text from Excel spreadsheets."""

    def __init__(self):
        """Initialize Excel processor."""
        pass

    def process(self, file_path: str) -> List[Tuple[str, dict]]:
        """
        Process Excel file and extract all text.

        Args:
            file_path: Path to Excel file

        Returns:
            List of (text, metadata) tuples
        """
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            results = []

            logger.info(f"Processing Excel file with {len(workbook.sheetnames)} sheets")

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Extract all cell values
                text_parts = []
                row_count = 0
                cell_count = 0

                for row in sheet.iter_rows(values_only=True):
                    row_count += 1
                    row_values = []

                    for cell_value in row:
                        if cell_value is not None:
                            cell_count += 1
                            # Convert to string
                            cell_text = str(cell_value).strip()
                            if cell_text:
                                row_values.append(cell_text)

                    if row_values:
                        # Join cells in row with tab separator
                        text_parts.append("\t".join(row_values))

                # Combine all rows
                full_text = "\n".join(text_parts)

                metadata = {
                    "source": Path(file_path).name,
                    "sheet_name": sheet_name,
                    "rows": row_count,
                    "cells_with_data": cell_count,
                    "type": "excel",
                }

                if full_text.strip():
                    results.append((full_text, metadata))

            workbook.close()
            logger.info(f"Extracted data from {len(results)} sheets")

            return results

        except Exception as e:
            logger.error(f"Error processing Excel file: {e}")
            raise

    def process_with_structure(self, file_path: str) -> List[Tuple[str, dict]]:
        """
        Process Excel file preserving structure (headers, columns).

        Args:
            file_path: Path to Excel file

        Returns:
            List of (text, metadata) tuples with column information
        """
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            results = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Assume first row is header
                rows = list(sheet.iter_rows(values_only=True))

                if not rows:
                    continue

                headers = [str(h) if h is not None else f"Column_{i}" for i, h in enumerate(rows[0])]

                # Process each row
                text_parts = []
                text_parts.append("Headers: " + ", ".join(headers))
                text_parts.append("")

                for row_idx, row in enumerate(rows[1:], start=2):
                    row_text = []

                    for col_idx, (header, value) in enumerate(zip(headers, row)):
                        if value is not None:
                            value_str = str(value).strip()
                            if value_str:
                                row_text.append(f"{header}: {value_str}")

                    if row_text:
                        text_parts.append(f"Row {row_idx}: " + " | ".join(row_text))

                full_text = "\n".join(text_parts)

                metadata = {
                    "source": Path(file_path).name,
                    "sheet_name": sheet_name,
                    "columns": headers,
                    "row_count": len(rows) - 1,  # Exclude header
                    "type": "excel_structured",
                }

                if full_text.strip():
                    results.append((full_text, metadata))

            workbook.close()
            return results

        except Exception as e:
            logger.error(f"Error processing Excel file with structure: {e}")
            raise
