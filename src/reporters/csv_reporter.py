"""CSV and Excel export for scan results."""

import csv
from pathlib import Path
from typing import Optional, Dict
from datetime import datetime

from src.models import ScanResult
from src.utils import get_logger

logger = get_logger(__name__)


class CSVReporter:
    """Export scan results to CSV format."""

    def export_findings(self, scan_result: ScanResult, output_path: str) -> None:
        """
        Export findings to CSV.

        Args:
            scan_result: Scan results
            output_path: Output CSV file path
        """
        try:
            with open(output_path, "w", newline="", encoding="utf-8") as f:
                if not scan_result.findings:
                    logger.warning("No findings to export")
                    return

                # Define CSV columns
                fieldnames = [
                    "id",
                    "source",
                    "location",
                    "pii_type",
                    "context",
                    "confidence",
                    "severity",
                    "gdpr_articles",
                    "recommendation",
                    "timestamp",
                    "line_number",
                    "page_number",
                ]

                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()

                for finding in scan_result.findings:
                    row = {
                        "id": finding.id,
                        "source": finding.source,
                        "location": finding.location,
                        "pii_type": finding.pii_type,
                        "context": finding.context[:100] + "..." if len(finding.context) > 100 else finding.context,
                        "confidence": f"{finding.confidence:.2f}",
                        "severity": finding.severity,
                        "gdpr_articles": ", ".join(finding.gdpr_articles),
                        "recommendation": finding.recommendation,
                        "timestamp": finding.timestamp.isoformat() if finding.timestamp else "",
                        "line_number": finding.line_number or "",
                        "page_number": finding.page_number or "",
                    }
                    writer.writerow(row)

            logger.info(f"CSV export complete: {output_path}")

        except Exception as e:
            logger.error(f"CSV export failed: {e}")
            raise

    def export_summary(self, scan_result: ScanResult, output_path: str) -> None:
        """
        Export summary statistics to CSV.

        Args:
            scan_result: Scan results
            output_path: Output CSV file path
        """
        try:
            with open(output_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)

                # Header
                writer.writerow(["Privacy Scan Summary"])
                writer.writerow(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
                writer.writerow([])

                # Basic info
                writer.writerow(["Scan Information"])
                writer.writerow(["Source", scan_result.source])
                writer.writerow(["Type", scan_result.source_type])
                writer.writerow(["Status", scan_result.status])
                writer.writerow(["Duration (seconds)", f"{scan_result.duration_seconds:.2f}"])
                writer.writerow(["Total Findings", scan_result.total_findings])
                writer.writerow([])

                # By severity
                writer.writerow(["Findings by Severity"])
                writer.writerow(["Severity", "Count"])
                for severity, count in sorted(scan_result.findings_by_severity.items()):
                    writer.writerow([severity, count])
                writer.writerow([])

                # By type
                writer.writerow(["Findings by PII Type"])
                writer.writerow(["PII Type", "Count"])
                for pii_type, count in sorted(
                    scan_result.findings_by_type.items(), key=lambda x: x[1], reverse=True
                ):
                    writer.writerow([pii_type, count])

            logger.info(f"CSV summary export complete: {output_path}")

        except Exception as e:
            logger.error(f"CSV summary export failed: {e}")
            raise


class ExcelReporter:
    """Export scan results to Excel format."""

    def export_findings(self, scan_result: ScanResult, output_path: str) -> None:
        """
        Export findings to Excel with multiple sheets.

        Args:
            scan_result: Scan results
            output_path: Output Excel file path
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

        except ImportError:
            logger.error("openpyxl required for Excel export. Install with: pip install openpyxl")
            # Fallback to CSV
            csv_path = output_path.replace(".xlsx", ".csv")
            logger.info(f"Falling back to CSV export: {csv_path}")
            csv_reporter = CSVReporter()
            csv_reporter.export_findings(scan_result, csv_path)
            return

        try:
            wb = openpyxl.Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Sheet 1: Summary
            ws_summary = wb.create_sheet("Summary")
            self._create_summary_sheet(ws_summary, scan_result)

            # Sheet 2: All Findings
            ws_findings = wb.create_sheet("Findings")
            self._create_findings_sheet(ws_findings, scan_result)

            # Sheet 3: By Severity
            ws_severity = wb.create_sheet("By Severity")
            self._create_severity_sheet(ws_severity, scan_result)

            # Sheet 4: By Type
            ws_type = wb.create_sheet("By Type")
            self._create_type_sheet(ws_type, scan_result)

            # Save workbook
            wb.save(output_path)
            logger.info(f"Excel export complete: {output_path}")

        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            raise

    def _create_summary_sheet(self, ws, scan_result: ScanResult) -> None:
        """Create summary sheet."""
        from openpyxl.styles import Font, PatternFill

        # Title
        ws["A1"] = "Privacy Scan Summary"
        ws["A1"].font = Font(size=16, bold=True)

        # Info
        row = 3
        info_data = [
            ["Source", scan_result.source],
            ["Type", scan_result.source_type],
            ["Status", scan_result.status],
            ["Duration (s)", f"{scan_result.duration_seconds:.2f}"],
            ["Total Findings", scan_result.total_findings],
            ["Scan ID", scan_result.scan_id],
        ]

        for label, value in info_data:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = value
            row += 1

    def _create_findings_sheet(self, ws, scan_result: ScanResult) -> None:
        """Create findings sheet."""
        from openpyxl.styles import Font, PatternFill

        # Headers
        headers = [
            "ID",
            "Source",
            "Location",
            "PII Type",
            "Severity",
            "Confidence",
            "Context",
            "GDPR Articles",
            "Recommendation",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Data
        for row, finding in enumerate(scan_result.findings, 2):
            ws.cell(row, 1, finding.id)
            ws.cell(row, 2, finding.source)
            ws.cell(row, 3, finding.location)
            ws.cell(row, 4, finding.pii_type)
            ws.cell(row, 5, finding.severity)
            ws.cell(row, 6, f"{finding.confidence:.2f}")
            ws.cell(row, 7, finding.context[:100])
            ws.cell(row, 8, ", ".join(finding.gdpr_articles))
            ws.cell(row, 9, finding.recommendation)

        # Adjust column widths
        for col in range(1, 10):
            ws.column_dimensions[chr(64 + col)].width = 20

    def _create_severity_sheet(self, ws, scan_result: ScanResult) -> None:
        """Create severity breakdown sheet."""
        from openpyxl.styles import Font

        ws["A1"] = "Severity"
        ws["A1"].font = Font(bold=True)
        ws["B1"] = "Count"
        ws["B1"].font = Font(bold=True)

        row = 2
        for severity, count in sorted(scan_result.findings_by_severity.items()):
            ws[f"A{row}"] = severity
            ws[f"B{row}"] = count
            row += 1

    def _create_type_sheet(self, ws, scan_result: ScanResult) -> None:
        """Create PII type breakdown sheet."""
        from openpyxl.styles import Font

        ws["A1"] = "PII Type"
        ws["A1"].font = Font(bold=True)
        ws["B1"] = "Count"
        ws["B1"].font = Font(bold=True)

        row = 2
        for pii_type, count in sorted(
            scan_result.findings_by_type.items(), key=lambda x: x[1], reverse=True
        ):
            ws[f"A{row}"] = pii_type
            ws[f"B{row}"] = count
            row += 1
